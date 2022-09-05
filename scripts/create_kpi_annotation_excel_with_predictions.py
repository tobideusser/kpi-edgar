import os
import pickle
import random
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm

from edgar.data_classes import Corpus, Document, Paragraph, Sentence, Relation


def add_kpi_entry(kpi_dict, entity, num_kpi, sentence):
    kpi_dict["kpi"][entity.get_words(sentence.words)[0].id_] = {"num": num_kpi, "davon": []}


def add_davon_entry(kpi_dict, kpi_entity, davon_entity, sentence):
    kpi_dict["davon"][davon_entity.get_words(sentence.words)[0].id_] = {"kpi": kpi_entity}
    kpi_dict["kpi"][kpi_entity.get_words(sentence.words)[0].id_]["davon"].append(
        davon_entity.get_words(sentence.words)[0].id_
    )


def lookup_kpi(kpi_dict, entity, sentence):
    return (
        kpi_dict["kpi"][entity.get_words(sentence.words)[0].id_]["num"],
        kpi_dict["kpi"][entity.get_words(sentence.words)[0].id_]["davon"],
    )


def lookup_highest_unused_kpi_number(kpi_dict):
    kpi_nums = [attr["num"] for kpi, attr in kpi_dict["kpi"].items()]
    if len(kpi_nums):
        return max(kpi_nums) + 1, []
    else:
        return 1, []


def lookup_davon(kpi_dict, entity, sentence):
    if entity.get_words(sentence.words)[0].id_ in kpi_dict["davon"]:
        return kpi_dict["davon"][entity.get_words(sentence.words)[0].id_]["kpi"]


def obtain_updated_entity(tokens, new_type):
    return {"tokens": tokens, "type": new_type}


def _convert_relations_to_excel_notation(
    sentence: Sentence, type_: str, relations: Optional[List[Relation]] = None, add_unrelated_entities: bool = False
) -> List[Optional[str]]:

    updated_entities = []
    kpi_dict = {"kpi": {}, "davon": {}}
    entity_start_ids = set()
    if relations:
        for relation in relations:
            head_entity = relation.head_entity
            tail_entity = relation.tail_entity
            num_kpi = 0
            if add_unrelated_entities:
                entity_start_ids.add(head_entity.start)
                entity_start_ids.add(tail_entity.start)

            if head_entity.type_ == "kpi" and tail_entity.type_ != "davon":
                if head_entity.get_words(sentence.words)[0].id_ not in kpi_dict["kpi"]:
                    num_kpi += 1
                    add_kpi_entry(kpi_dict, head_entity, num_kpi, sentence)
            elif head_entity.type_ != "davon" and tail_entity.type_ == "kpi":
                if tail_entity.get_words(sentence.words)[0].id_ not in kpi_dict["kpi"]:
                    num_kpi += 1
                    add_kpi_entry(kpi_dict, tail_entity, num_kpi, sentence)
            elif head_entity.type_ == "kpi" and tail_entity.type_ == "davon":
                if head_entity.get_words(sentence.words)[0].id_ not in kpi_dict["kpi"]:
                    num_kpi += 1
                    add_kpi_entry(kpi_dict, head_entity, num_kpi, sentence)
                add_davon_entry(kpi_dict, head_entity, tail_entity, sentence)

            elif head_entity.type_ == "davon" and tail_entity.type_ == "kpi":
                if tail_entity.get_words(sentence.words)[0].id_ not in kpi_dict["kpi"]:
                    num_kpi += 1
                    add_kpi_entry(kpi_dict, tail_entity, num_kpi, sentence)
                add_davon_entry(kpi_dict, tail_entity, head_entity, sentence)

        for relation in relations:
            head_entity = relation.head_entity
            tail_entity = relation.tail_entity

            if head_entity.type_ == "kpi" and tail_entity.type_ != "davon":
                num, _ = lookup_kpi(kpi_dict, head_entity, sentence)
                updated_head_entity = obtain_updated_entity(
                    head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                )
                updated_tail_entity = obtain_updated_entity(
                    tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                )

            elif head_entity.type_ != "davon" and tail_entity.type_ == "kpi":
                num, _ = lookup_kpi(kpi_dict, tail_entity, sentence)
                updated_head_entity = obtain_updated_entity(
                    head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                )
                updated_tail_entity = obtain_updated_entity(
                    tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                )

            elif head_entity.type_ == "davon" and tail_entity.type_ == "kpi":
                num, davon_list = lookup_kpi(kpi_dict, tail_entity, sentence)
                if len(davon_list) == 1:
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                    )
                else:
                    davon_index = davon_list.index(head_entity.get_words(sentence.words)[0].id_)
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}_{davon_index + 1}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                    )

            elif head_entity.type_ == "kpi" and tail_entity.type_ == "davon":
                num, davon_list = lookup_kpi(kpi_dict, head_entity, sentence)
                if len(davon_list) == 1:
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                    )
                else:
                    davon_index = davon_list.index(tail_entity.get_words(sentence.words)[0].id_)
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}_{davon_index + 1}"
                    )

            elif head_entity.type_ != "kpi" and tail_entity.type_ == "davon":
                kpi = lookup_davon(kpi_dict, tail_entity, sentence)
                if kpi is not None:
                    num, davon_list = lookup_kpi(kpi_dict, kpi, sentence)
                else:
                    num, davon_list = lookup_highest_unused_kpi_number(kpi_dict)
                if len(davon_list) <= 1:
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                    )
                else:
                    davon_index = davon_list.index(tail_entity.get_words(sentence.words)[0].id_)
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}_{davon_index + 1}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}_{davon_index + 1}"
                    )

            elif head_entity.type_ == "davon" and tail_entity.type_ != "kpi":
                kpi = lookup_davon(kpi_dict, head_entity, sentence)
                if kpi is not None:
                    num, davon_list = lookup_kpi(kpi_dict, kpi, sentence)
                else:
                    num, davon_list = lookup_highest_unused_kpi_number(kpi_dict)
                if len(davon_list) <= 1:
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}"
                    )
                else:
                    davon_index = davon_list.index(head_entity.get_words(sentence.words)[0].id_)
                    updated_head_entity = obtain_updated_entity(
                        head_entity.get_words(sentence.words), head_entity.type_ + f"_{num}_{davon_index + 1}"
                    )
                    updated_tail_entity = obtain_updated_entity(
                        tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{num}_{davon_index + 1}"
                    )
            else:
                updated_head_entity = obtain_updated_entity(
                    head_entity.get_words(sentence.words), head_entity.type_ + f"_{1}"
                )
                updated_tail_entity = obtain_updated_entity(
                    tail_entity.get_words(sentence.words), tail_entity.type_ + f"_{1}"
                )
            updated_entities.extend([updated_head_entity, updated_tail_entity])

    converted_relations = [None] * sentence.n_words
    if add_unrelated_entities:
        if type_ == "anno":
            if sentence.entities_anno:
                for entity in sentence.entities_anno:
                    if entity.start not in entity_start_ids:
                        updated_entities.append(
                            {
                                "tokens": entity.get_words(sentence.words),
                                "type": entity.type_ + "_0" if entity.type_ != "false_positive" else entity.type_,
                            }
                        )
        elif type_ == "pred":
            if sentence.entities_pred:
                for entity in sentence.entities_pred:
                    if entity.start not in entity_start_ids:
                        updated_entities.append(
                            {
                                "tokens": entity.get_words(sentence.words),
                                "type": entity.type_ + "_0" if entity.type_ != "false_positive" else entity.type_,
                            }
                        )
    for entity in updated_entities:
        for token in entity["tokens"]:
            converted_relations[token.id_] = entity["type"]

    return converted_relations


def _write_document_name(ws: Worksheet, document: Document, new_split_type: Optional[str] = None) -> Worksheet:
    # get split type
    split_type = None
    for segment in document:
        if isinstance(segment, Paragraph):
            for sentence in segment:
                split_type = sentence.split_type
                break
        if split_type:
            break
    if new_split_type:
        cells = [WriteOnlyCell(ws, value=document.id_), WriteOnlyCell(ws, value=new_split_type)]
    else:
        cells = [WriteOnlyCell(ws, value=document.id_), WriteOnlyCell(ws, value=split_type)]
    row = []
    for i, cell in enumerate(cells):
        if i == 0:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        row.append(cell)
    ws.append(row)
    ws.append([])
    return ws


def _write_segment_id(ws: Worksheet, segment_id: int) -> Worksheet:
    cells = [WriteOnlyCell(ws, value="Blob ID"), WriteOnlyCell(ws, value=segment_id)]
    row = []
    for cell in cells:
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        row.append(cell)
    ws.append(row)
    return ws


def _write_sentence_id(ws: Worksheet, sentence_id: int) -> Worksheet:
    cells = [WriteOnlyCell(ws, value="Sentence ID"), WriteOnlyCell(ws, value=sentence_id)]
    row = []
    for cell in cells:
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        row.append(cell)
    ws.append(row)
    return ws


def _write_sentence(
    ws: Worksheet,
    sentence: Sentence,
    converted_relations_pred: List[Optional[str]],
    converted_relations_anno: List[Optional[str]],
) -> Worksheet:
    cell = WriteOnlyCell(ws, value="Sentence")
    cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
    row = [cell]
    for word, pred, anno in zip(sentence.words, converted_relations_pred, converted_relations_anno):
        cell = WriteOnlyCell(ws, value=word.value)
        cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        if pred != anno:  # and anno != "false_positive" and (anno is None or anno[-1] != "0"):
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(name="Calibri", bold=True)
        else:
            cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
        row.append(cell)
    ws.append(row)
    return ws


def _get_relation_tag_color_code(tag: str):
    if tag[:2] == "cy":
        return "fff200"
    elif tag[:3] == "kpi" and tag[4] != "c":
        return "72bf44"
    elif tag[:3] == "kpi" and tag[4] == "c":
        return "00aaad"
    elif tag[:3] == "py_":
        return "faa61a"
    elif tag[:3] == "py1":
        return "f58220"
    elif tag[:3] == "inc":
        return "bd7cb5"
    elif tag[:3] == "dec":
        return "ed1c24"
    elif len(tag) > 4 and tag[:4] == "attr":
        return "7da7d8"
    elif len(tag) > 6 and tag[:5] == "davon" and tag[6] in [str(num) for num in range(10)]:
        return "c2e0ae"
    elif len(tag) > 7 and tag[:5] == "davon" and tag[6:8] == "cy":
        return "fff9ae"
    elif len(tag) > 7 and tag[:5] == "davon" and tag[6:8] == "py":
        return "ffdaa2"
    elif len(tag) > 7 and tag[:5] == "davon" and tag[6:8] == "in":
        return "c7a0cb"
    elif len(tag) > 7 and tag[:5] == "davon" and tag[6:8] == "de":
        return "f7a19a"
    elif len(tag) > 7 and tag[:5] == "davon" and tag[6:8] == "co":
        return "948a54"
    elif len(tag) > 5 and tag[:5] == "false":
        return "000000"
    else:
        return "DCDCDC"


def _write_converted_relations(ws: Worksheet, converted_relations: List[Optional[str]], row_name: str) -> Worksheet:
    cell = WriteOnlyCell(ws, value=row_name)
    cell.fill = PatternFill(start_color="DCDCDC", end_color="DCDCDC", fill_type="solid")
    row = [cell]
    for tag in converted_relations:
        cell = WriteOnlyCell(ws, value=tag)
        color_code = _get_relation_tag_color_code(tag) if tag else "DCDCDC"
        cell.fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        if color_code == "000000":
            cell.font = Font(color="FFFFFF", name="Calibri")
        row.append(cell)
    ws.append(row)
    return ws


def main():

    len_test_valid_set = 8
    # todo: annotate documents, then run this again?
    predicted_corpus_dir = "/scratch/data/edgar/above200B/DataTagging/000/"
    predicted_corpus_path = os.path.join(predicted_corpus_dir, "corpus_tagged.p")
    corpus = pickle.load(open(predicted_corpus_path, "rb"))
    save_path = "/scratch/data/edgar/"

    corpus = Corpus.from_dict(corpus)

    wb = Workbook(write_only=True)

    test_set = []
    valid_set = []
    all_other_docs = []

    for document in corpus.documents:

        split_type = ""
        for segment in document.segments:
            if len(segment.sentences) > 0:
                split_type = segment.sentences[0].split_type
                break

        if split_type == "test":
            test_set.append(document)
        elif split_type == "valid":
            valid_set.append(document)
        else:
            all_other_docs.append(document)

    random.seed(1337)

    ordered_docs = []
    random.shuffle(all_other_docs)

    # add valid set
    ordered_docs.extend(valid_set)
    # add from all_other_docs until len_test_valid_set is reached
    ordered_docs.extend(all_other_docs[: (len_test_valid_set - len(valid_set))])
    del all_other_docs[: (len_test_valid_set - len(valid_set))]

    # add test set
    ordered_docs.extend(test_set)
    ordered_docs.extend(all_other_docs)

    for sheet_number, document in enumerate(tqdm(ordered_docs, desc="Writing document to xlsx"), 0):
        if sheet_number < len_test_valid_set:
            split_type = "valid"
        elif sheet_number < 2 * len_test_valid_set:
            split_type = "test"
        else:
            split_type = "train"

        sheet_name = str(sheet_number)
        ws = wb.create_sheet(sheet_name)
        ws = _write_document_name(ws, document, new_split_type=split_type)

        for segment in document.segments:
            if isinstance(segment, Paragraph):
                for sentence in segment.sentences:

                    converted_relations_pred = _convert_relations_to_excel_notation(
                        sentence=sentence, relations=sentence.relations_pred, type_="pred", add_unrelated_entities=True
                    )
                    converted_relations_anno = _convert_relations_to_excel_notation(
                        sentence=sentence, relations=sentence.relations_anno, type_="anno", add_unrelated_entities=True
                    )

                    ws = _write_segment_id(ws, segment.id_)
                    ws = _write_sentence_id(ws, sentence.id_)
                    ws = _write_sentence(ws, sentence, converted_relations_pred, converted_relations_anno)
                    ws = _write_converted_relations(ws, converted_relations_pred, "Auto Annotations")
                    ws = _write_converted_relations(ws, converted_relations_anno, "Manual Annotations")

                    ws.append([WriteOnlyCell(ws, value="Comment")])
                    ws.append([])

    wb.save(filename=os.path.join(save_path, "edgar_relation_annotations.xlsx"))


if __name__ == "__main__":
    main()
